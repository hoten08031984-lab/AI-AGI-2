import http.server
import socketserver
import os
import time
import json
import sys
import threading
import unicodedata
import logging
import subprocess
import openpyxl
import shutil

# --- Logging: ghi ra cả console và file để debug khi chạy pythonw ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_PATH = os.path.join(BASE_DIR, 'server.log')

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_PATH, encoding='utf-8', mode='w'),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger('dashboard')

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return unicodedata.normalize('NFC', s)

def normalize_loai_cp(val):
    s = clean_str(val)
    if not s or s.lower() == 'none':
        return 'Chưa phân loại'
    s_lower = s.lower()
    # Unify all typos & accent variants of Hành chính phí
    if 'hành' in s_lower or 'hanh' in s_lower or 'hà' in s_lower and 'chí' in s_lower or 'hà' in s_lower and 'chi' in s_lower:
        if 'chí' in s_lower or 'chi' in s_lower:
            return 'Hành chính phí'
    if 'công tác' in s_lower or 'cong tac' in s_lower or 'cô' in s_lower and 'tá' in s_lower:
        return 'Công tác phí'
    if 'bảo trì' in s_lower or 'bao tri' in s_lower:
        return 'Phí bảo trì'
    return s

def format_so_hd(val):
    s = clean_str(val)
    if not s:
        return ''
    if s.endswith('.0'):
        return s[:-2]
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except:
        pass
    return s

def format_ngay_hd(val):
    s = clean_str(val)[:10]
    if not s or s.lower() == 'none':
        return ''
    import re
    m_iso = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m_iso:
        return f"{m_iso.group(3).zfill(2)}-{m_iso.group(2).zfill(2)}-{m_iso.group(1)}"
    m_mdy = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})', s)
    if m_mdy:
        p1, p2, yyyy = int(m_mdy.group(1)), int(m_mdy.group(2)), m_mdy.group(3)
        if p1 > 12:
            dd, mm = str(p1).zfill(2), str(p2).zfill(2)
        elif p2 > 12:
            mm, dd = str(p1).zfill(2), str(p2).zfill(2)
        else:
            mm, dd = str(p1).zfill(2), str(p2).zfill(2)
        return f"{dd}-{mm}-{yyyy}"
    return s



PORT = 8080
DIRECTORY = BASE_DIR
EXCEL_FILENAME = "THEO DOI HOP DONG-2_Optimized.xlsx"
EXCEL_PATH = os.path.join(BASE_DIR, EXCEL_FILENAME)
JS_DATA_PATH = os.path.join(BASE_DIR, "dashboard_data.js")

def find_source_excel():
    """Tìm đường dẫn file Excel trên OneDrive ở các vị trí phổ biến."""
    candidates = [
        r"D:\OneDrive - SABECO\Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx",
        os.path.join(os.environ.get('OneDriveCommercial', ''), r"Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx"),
        os.path.join(os.environ.get('OneDrive', ''), r"Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx"),
        os.path.join(os.environ.get('USERPROFILE', ''), r"OneDrive - SABECO\Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx"),
        os.path.join(os.environ.get('USERPROFILE', ''), r"OneDrive\Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx"),
    ]
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return None

SOURCE_EXCEL_PATH = find_source_excel() or r"D:\OneDrive - SABECO\Công Việc\0 -THANH TOÁN THÁNG HD\THEO DOI HOP DONG-2_Optimized.xlsx"

last_mtime = 0
last_check_time = 0
extract_lock = threading.Lock()

def sync_from_source_excel(force=False):
    """Tự động đồng bộ và chép đè file Excel từ đường dẫn OneDrive nếu tồn tại."""
    src = find_source_excel()
    if not src or not os.path.exists(src):
        log.warning(f"Không tìm thấy file nguồn OneDrive.")
        return False
    try:
        mtime_src = os.path.getmtime(src)
        mtime_dst = os.path.getmtime(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else 0

        if force or mtime_src > mtime_dst or not os.path.exists(EXCEL_PATH):
            tmp_path = EXCEL_PATH + ".tmp"
            copied = False

            # Try shutil.copy2 first
            try:
                shutil.copy2(src, tmp_path)
                copied = True
            except Exception:
                pass

            # Fallback to binary stream
            if not copied:
                try:
                    with open(src, 'rb') as f_src:
                        data = f_src.read()
                    with open(tmp_path, 'wb') as f_dst:
                        f_dst.write(data)
                    copied = True
                except Exception as e_bin:
                    log.warning(f"Không thể đọc file từ OneDrive: {e_bin}")

            if os.path.exists(tmp_path) and os.path.getsize(tmp_path) > 10000:
                if os.path.exists(EXCEL_PATH):
                    try: os.remove(EXCEL_PATH)
                    except: pass
                os.replace(tmp_path, EXCEL_PATH)
                log.info(f"Đã đồng bộ chép đè file Excel mới nhất từ OneDrive:\n  -> {src}")
                return True
            elif os.path.exists(tmp_path):
                os.remove(tmp_path)
    except Exception as e:
        log.warning(f"Lỗi khi chép đè file Excel từ OneDrive: {e}")
    return False

def read_excel_via_powershell(abs_path):
    """Đọc dữ liệu Excel qua PowerShell COM Bridge (hỗ trợ 100% file mã hóa DRM / Sensitivity Label)."""
    import subprocess
    
    full_path = os.path.abspath(abs_path).replace("'", "''")
    json_out = os.path.join(BASE_DIR, "_temp_extracted_data.json").replace("'", "''")
    ps_script = f"""
$ErrorActionPreference = 'SilentlyContinue'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$excel.ScreenUpdating = $false
$excel.AutomationSecurity = 1
$wb = $excel.Workbooks.Open('{full_path}', 0, $true)
Start-Sleep -Milliseconds 800
$ws = $wb.Sheets.Item('CHI')
$usedRange = $ws.UsedRange
$data = $usedRange.Value2
$wb.Close($false)
$excel.Quit()
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null

$rows = @()
$maxR = $data.GetLength(0)
$maxC = $data.GetLength(1)
for ($r = 1; $r -le $maxR; $r++) {{
    $row = @()
    for ($c = 1; $c -le $maxC; $c++) {{
        $val = $data[$r, $c]
        $row += $val
    }}
    $rows += ,$row
}}
$json = $rows | ConvertTo-Json -Depth 5 -Compress
[System.IO.File]::WriteAllText('{json_out}', $json, [System.Text.Encoding]::UTF8)
"""
    try:
        cmd = ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script]
        proc = subprocess.run(cmd, capture_output=True, timeout=30)
        if os.path.exists(json_out):
            with open(json_out, 'r', encoding='utf-8-sig') as f:
                vals = json.load(f)
            try: os.remove(json_out)
            except: pass
            if vals and len(vals) > 0:
                log.info(f"  -> Đọc thành công qua PowerShell COM Bridge ({len(vals)} dòng).")
                return vals
    except Exception as e_ps:
        log.warning(f"  -> PowerShell COM Bridge gặp lỗi: {e_ps}")
        if os.path.exists(json_out):
            try: os.remove(json_out)
            except: pass
    return None

def extract_excel_data(force=False):
    global last_mtime, last_check_time
    
    now = time.time()
    if not force and now - last_check_time < 2:
        return False

    if not extract_lock.acquire(blocking=False):
        return False
        
    try:
        last_check_time = now
        sync_from_source_excel(force=force)

        if not os.path.exists(EXCEL_PATH):
            print(f"File Excel not found: {EXCEL_PATH}")
            return False
        
        current_mtime = os.path.getmtime(EXCEL_PATH)
        if not force and current_mtime <= last_mtime:
            return False  # No change

        log.info(f"Phát hiện kiểm tra Excel (Force={force})! Đang đọc và cập nhật dữ liệu...")
        
        vals = None
        abs_src = os.path.abspath(EXCEL_PATH)

        # 1. Strategy 1: openpyxl (nhanh, chuẩn cho file .xlsx thông thường)
        try:
            wb_ox = openpyxl.load_workbook(abs_src, data_only=True)
            if 'CHI' in wb_ox.sheetnames:
                ws_ox = wb_ox['CHI']
                vals = [list(row) for row in ws_ox.iter_rows(values_only=True)]
                log.info(f"  -> Đọc thành công bằng openpyxl.")
            wb_ox.close()
        except Exception as e_ox:
            log.info(f"  -> openpyxl không đọc được ({e_ox}), chuyển sang PowerShell COM Bridge...")

        # 2. Strategy 2: PowerShell COM Bridge (tương thích 100% DRM / Sensitivity Label của SABECO)
        if not vals:
            vals = read_excel_via_powershell(abs_src)

        # 3. Strategy 3: win32com fallback
        if not vals:
            try:
                import pythoncom
                import win32com.client as win32

                pythoncom.CoInitialize()
                _excel = win32.Dispatch('Excel.Application')
                _excel.Visible = False
                _excel.DisplayAlerts = False
                _excel.AutomationSecurity = 1
                try:
                    _wb = _excel.Workbooks.Open(abs_src, 0, True)
                    time.sleep(1)
                    ws = _wb.Sheets('CHI')
                    vals = ws.UsedRange.Value
                    log.info(f"  -> Đọc thành công bằng win32com ({len(vals) if vals else 0} dòng).")
                    try: _wb.Close(False)
                    except: pass
                finally:
                    try: _excel.Quit()
                    except: pass
                    pythoncom.CoUninitialize()
            except Exception as e_win32:
                log.error(f"  -> win32com không đọc được dữ liệu: {e_win32}")

        if not vals:
            log.error(f"  -> Không đọc được dữ liệu Excel từ bất kỳ phương thức nào.")
            return False

        clean_rows = []
        for r in vals:
            if isinstance(r, dict) and 'value' in r:
                clean_rows.append(r['value'])
            elif isinstance(r, (list, tuple)):
                clean_rows.append(list(r))

        rows = [r for r in clean_rows if r and any(x is not None for x in r)]
        data = rows[1:] if len(rows) > 0 else []
            
        processed_items = []
        for idx, r in enumerate(data):
            if not r or not any(x is not None for x in r):
                continue

            r_padded = list(r) + [None] * (17 - len(r))

            raw_loai_cp = r_padded[0]
            loai_cp = normalize_loai_cp(raw_loai_cp)

            tieu_muc = clean_str(r_padded[1])
            so_hd = format_so_hd(r_padded[2])
            ngay_hd = format_ngay_hd(r_padded[3])
            thang = clean_str(r_padded[4])
            ly_do = clean_str(r_padded[5])
            chi_tiet = clean_str(r_padded[6])
            chi_tiet_hd = clean_str(r_padded[7])
            kho = clean_str(r_padded[8])

            try: st_no_vat = float(r_padded[9]) if r_padded[9] is not None else 0.0
            except: st_no_vat = 0.0

            try: vat = float(r_padded[10]) if r_padded[10] is not None else 0.0
            except: vat = 0.0

            try: st_vat = float(r_padded[11]) if r_padded[11] is not None else 0.0
            except: st_vat = 0.0

            ngay_tt = clean_str(r_padded[12])[:10]
            nguoi_thu_huong = clean_str(r_padded[13])
            stk = clean_str(r_padded[14])
            ngan_hang = clean_str(r_padded[15])

            try:
                parsed_nam = int(float(r_padded[16])) if r_padded[16] is not None else 0
                if parsed_nam <= 1900 or parsed_nam > 2100:
                    nam = 'N/A'
                else:
                    nam = parsed_nam
            except:
                nam = 'N/A'

            # Ignore only completely empty noise rows
            if loai_cp == 'Chưa phân loại' and not tieu_muc and not so_hd and not ly_do and not chi_tiet and st_vat == 0 and st_no_vat == 0:
                continue

            processed_items.append({
                'id': len(processed_items) + 1,
                'loai_cp': loai_cp,
                'tieu_muc': tieu_muc,
                'so_hd': so_hd,
                'ngay_hd': ngay_hd,
                'thang': thang,
                'ly_do': ly_do,
                'chi_tiet': chi_tiet,
                'chi_tiet_hd': chi_tiet_hd,
                'kho': kho if kho and kho != 'None' else 'Khác',
                'st_no_vat': st_no_vat,
                'vat': vat,
                'st_vat': st_vat,
                'ngay_tt': ngay_tt,
                'nguoi_thu_huong': nguoi_thu_huong if nguoi_thu_huong and nguoi_thu_huong != 'None' else '',
                'stk': stk if stk and stk != 'None' else '',
                'ngan_hang': ngan_hang if ngan_hang and ngan_hang != 'None' else '',
                'nam': nam
            })
            
        sync_info = {
            'last_updated': time.strftime('%Y-%m-%d %H:%M:%S'),
            'total_rows': len(processed_items)
        }
        
        with open(JS_DATA_PATH, 'w', encoding='utf-8') as f:
            f.write(f'window.SYNC_INFO = {json.dumps(sync_info, ensure_ascii=False)};\n')
            f.write('window.RAW_DATA = ' + json.dumps(processed_items, ensure_ascii=False) + ';')

        last_mtime = current_mtime
        log.info(f"Cập nhật thành công! Tổng số {len(processed_items)} dòng.")
        return True
    except Exception as e:
        log.error(f"Lỗi khi đọc file Excel: {e}", exc_info=True)
        return False
    finally:
        extract_lock.release()


class AutoUpdateHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def do_GET(self):
        try:
            if self.path.startswith('/api/refresh'):
                success = extract_excel_data(force=True)
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                res = json.dumps({"status": "ok", "success": success}).encode('utf-8')
                self.wfile.write(res)
                return
            elif self.path in ('/', '/index.html') or self.path.startswith('/index.html'):
                extract_excel_data(force=True)
        except Exception as e:
            log.error(f"Lỗi trong do_GET: {e}")
        return super().do_GET()

def kill_old_server_on_port(port):
    """Kill any existing process using the given port to avoid Address-already-in-use."""
    try:
        result = subprocess.run(
            ['netstat', '-ano'],
            capture_output=True, text=True, timeout=5
        )
        for line in result.stdout.splitlines():
            if f':{port}' in line and 'LISTENING' in line:
                parts = line.strip().split()
                pid = parts[-1]
                if pid.isdigit() and int(pid) != os.getpid():
                    log.info(f"Đang tắt server cũ (PID {pid}) trên port {port}...")
                    subprocess.run(['taskkill', '/F', '/PID', pid],
                                   capture_output=True, timeout=5)
                    time.sleep(0.5)
    except Exception as e:
        log.warning(f"Không thể kiểm tra port cũ: {e}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Dashboard Chi Phi Server')
    parser.add_argument('--sync-only', action='store_true', help='Chỉ đồng bộ file từ OneDrive và cập nhật dữ liệu JS rồi thoát')
    args = parser.parse_args()

    # Tắt server cũ trên port 8080 trước để giải phóng file lock (nếu server ngầm đang chạy)
    kill_old_server_on_port(PORT)

    if args.sync_only:
        log.info(f"=== Đang thực hiện đồng bộ file từ OneDrive ===")
        synced = sync_from_source_excel(force=True)
        if synced:
            log.info("[OK] Đã sao chép thành công file Excel mới nhất từ OneDrive!")
        else:
            log.info("[INFO] Khung đồng bộ: Không có thay đổi mới hoặc file bị khóa / không tìm thấy OneDrive.")
        
        log.info("Đang trích xuất dữ liệu sang Dashboard...")
        success = extract_excel_data(force=True)
        if success:
            log.info("[OK] Đã cập nhật xong dữ liệu Dashboard!")
        sys.exit(0 if (synced or success) else 0)

    log.info(f"=== Dashboard Server khởi động ===")
    log.info(f"Thư mục: {BASE_DIR}")
    log.info(f"File Excel: {EXCEL_PATH}")
    log.info(f"File Excel tồn tại: {os.path.exists(EXCEL_PATH)}")

    # Kill server cũ nếu còn chạy trên port
    kill_old_server_on_port(PORT)

    # Đọc Excel ở luồng phụ lúc khởi động để không chặn HTTP server lắng nghe
    def bg_extract():
        try:
            extract_excel_data()
        except Exception as e:
            log.warning(f"Không thể đọc Excel lúc khởi động: {e}. Đang dùng dữ liệu đã cache...")

    threading.Thread(target=bg_extract, daemon=True).start()

    socketserver.TCPServer.allow_reuse_address = True
    try:
        with socketserver.TCPServer(("", PORT), AutoUpdateHandler) as httpd:
            log.info(f"Server đang chạy tại http://localhost:{PORT}")
            log.info("Mỗi khi bạn mở hoặc F5 lại trang Dashboard, server sẽ tự kiểm tra và đọc file Excel mới nhất!")
            httpd.serve_forever()
    except Exception as e:
        log.error(f"KHÔNG THỂ KHỞI ĐỘNG SERVER: {e}")
        log.error(f"Port {PORT} có thể đang bị chiếm. Hãy thử chạy lại hoặc kiểm tra tasklist.")
        time.sleep(5)
