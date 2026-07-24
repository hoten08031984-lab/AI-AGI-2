import http.server
import socketserver
import os
import time
import json
import sys
import threading
import unicodedata
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# CẤU HÌNH — Chỉ cần đổi các giá trị này
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PORT = 8080
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")          # Tên file Excel
SHEET_NAME = "Sheet1"                                       # Tên sheet cần đọc
JS_DATA_PATH = os.path.join(BASE_DIR, "dashboard_data.js")
DIRECTORY = BASE_DIR

last_mtime = 0
last_check_time = 0
extract_lock = threading.Lock()


def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return unicodedata.normalize('NFC', s)


def extract_excel_data(force=False):
    global last_mtime, last_check_time

    now = time.time()
    if not force and now - last_check_time < 2:
        return False

    if not extract_lock.acquire(blocking=False):
        return False

    try:
        last_check_time = now
        if not os.path.exists(EXCEL_PATH):
            print(f"File Excel không tìm thấy: {EXCEL_PATH}")
            return False

        current_mtime = os.path.getmtime(EXCEL_PATH)
        if not force and current_mtime <= last_mtime:
            return False

        print(f"[{time.strftime('%H:%M:%S')}] Đang đọc Excel (Force={force})...")

        vals = None
        abs_src = os.path.abspath(EXCEL_PATH)

        # Thử đọc qua win32com trước (hỗ trợ computed values)
        try:
            import pythoncom
            import win32com.client as win32

            for attempt in range(3):
                pythoncom.CoInitialize()
                excel = None
                wb = None
                try:
                    try:
                        excel = win32.GetActiveObject('Excel.Application')
                    except:
                        excel = win32.DispatchEx('Excel.Application')
                        excel.Visible = False
                        excel.DisplayAlerts = False

                    wb_found = None
                    for open_wb in excel.Workbooks:
                        if open_wb.FullName.lower() == abs_src.lower():
                            wb_found = open_wb
                            break

                    if wb_found:
                        wb = wb_found
                        close_wb = False
                    else:
                        wb = excel.Workbooks.Open(abs_src, 0, True)
                        close_wb = True

                    time.sleep(0.5)
                    ws = wb.Sheets(SHEET_NAME)
                    vals = ws.UsedRange.Value
                    if close_wb:
                        try: wb.Close(False)
                        except: pass
                    break
                except Exception as e:
                    print(f"Lần thử {attempt+1}: {e}")
                    time.sleep(1.0)
                finally:
                    pythoncom.CoUninitialize()

        except Exception as e_com:
            # Fallback: dùng openpyxl (không cần Excel cài sẵn)
            print(f"win32com không khả dụng ({e_com}), dùng openpyxl...")
            try:
                wb_ox = openpyxl.load_workbook(abs_src, data_only=True)
                if SHEET_NAME in wb_ox.sheetnames:
                    ws_ox = wb_ox[SHEET_NAME]
                    vals = [list(row) for row in ws_ox.iter_rows(values_only=True)]
            except Exception as e_ox:
                print(f"Lỗi openpyxl: {e_ox}")
                return False

        if not vals:
            return False

        # ============================================================
        # XỬ LÝ DỮ LIỆU — Tuỳ chỉnh theo cấu trúc file Excel của bạn
        # rows[0] = header, rows[1:] = data
        # ============================================================
        rows = [list(r) for r in vals if r and any(x is not None for x in r)]
        data_rows = rows[1:]

        processed = []
        for idx, r in enumerate(data_rows):
            r = list(r) + [None] * max(0, 10 - len(r))  # pad tới đủ cột

            # === ĐỔI MAPPING NÀY THEO CỘT THỰC TẾ CỦA FILE EXCEL ===
            item = {
                'id': idx + 1,
                'col_a': clean_str(r[0]),   # Cột A
                'col_b': clean_str(r[1]),   # Cột B
                'col_c': clean_str(r[2]),   # Cột C
                # ... thêm cột theo nhu cầu
            }

            # Bỏ qua dòng rỗng hoàn toàn
            if not any(item[k] for k in item if k != 'id'):
                continue

            processed.append(item)

        sync_info = {
            'last_updated': time.strftime('%Y-%m-%d %H:%M:%S'),
            'total_rows': len(processed)
        }

        with open(JS_DATA_PATH, 'w', encoding='utf-8') as f:
            f.write(f'window.SYNC_INFO = {json.dumps(sync_info, ensure_ascii=False)};\n')
            f.write('window.RAW_DATA = ' + json.dumps(processed, ensure_ascii=False) + ';')

        last_mtime = current_mtime
        print(f"[{time.strftime('%H:%M:%S')}] Xong! Đã xử lý {len(processed)} dòng.")
        return True

    except Exception as e:
        print(f"Lỗi: {e}")
        return False
    finally:
        extract_lock.release()


class AutoUpdateHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def end_headers(self):
        # Chống cache trình duyệt
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def do_GET(self):
        try:
            if self.path.startswith('/api/refresh'):
                # API endpoint để nút Làm Mới gọi
                success = extract_excel_data(force=True)
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "ok", "success": success}).encode())
                return
            elif self.path in ('/', '/index.html') or self.path.startswith('/index.html'):
                extract_excel_data(force=True)
        except Exception as e:
            print(f"Lỗi do_GET: {e}")
        return super().do_GET()

    def log_message(self, format, *args):
        # Giảm log noise — chỉ log request chính
        if args and ('200' in str(args[1]) or '304' in str(args[1])):
            pass
        super().log_message(format, *args)


if __name__ == '__main__':
    print("=" * 50)
    print(f"  Dashboard Server đang khởi động...")
    print(f"  Thư mục: {BASE_DIR}")
    print(f"  Excel:   {EXCEL_PATH}")
    print("=" * 50)

    try:
        extract_excel_data(force=True)
    except Exception as e:
        print(f"Không đọc được Excel lúc khởi động: {e}. Dùng dữ liệu cache...")

    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), AutoUpdateHandler) as httpd:
        print(f"\n✅ Server sẵn sàng tại: http://localhost:{PORT}")
        print("   Bấm Ctrl+C để dừng.\n")
        httpd.serve_forever()
